import streamlit as st
import base64
import openpyxl
import datetime
import time
import math

from openpyxl.styles.fonts import Font
from openpyxl.writer.excel import save_virtual_workbook


def main():
    st.title('Excelファイル分割')
    file = st.file_uploader('', type='xlsx',
                            accept_multiple_files=False, key=None, help=None)
    if file is not None:
        input_num = st.number_input('分割行数を指定してください(列は最終列まで自動取得されます)', min_value=1, max_value=None,
                                    value=100, step=1, format=None, key=None, help=None)
        st.write(f'{input_num}行ごとにファイルを分割します')
        input_text = st.text_input(
            '保存名を入力してください', max_chars=None, key=None, type='default', help=None)
        day_check = st.checkbox(
            '今日の日付をファイル名にする', value=False, key=None, help=None)
        if day_check:
            input_text = str(datetime.date.today())

        if len(input_text) > 1:
            st.write(f'保存後のファイル名：{input_text}_1.xlsx / {input_text}_2.xlsx...')
            form = st.form('form')
            submitted = form.form_submit_button("分割開始")
            if submitted:
                division(file, input_num, input_text)


def division(file, input_num, input_text):
    """ファイル分割"""
    # ワークブックを開く
    wb = openpyxl.load_workbook(file)
    # シートを開く
    ws = wb.worksheets[0]
    # スタート行を指定
    start_row = 1
    # 最終行を取得
    max_col = openpyxl.utils.get_column_letter(ws.max_column)
    # 分割行数をdiv_rowに代入
    div_row = input_num
    files = math.ceil(ws.max_row/div_row)
    for file_count in range(1, files+1):
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.worksheets[0]
        sheet_range = ws[f'A{start_row}':f'{max_col}{div_row}']
        data_list = load_cells(sheet_range)
        write_list(new_ws, data_list, 0, 1)
        new_ws = new_ws[f'A1':f'{max_col}{new_ws.max_row}']
        edit_font(new_ws)
        st.markdown(get_table_download_link(
            new_wb, input_text, file_count), unsafe_allow_html=True)
        start_row = div_row + 1
        div_row += input_num


def load_cells(sheet_range):
    data_list = [[]]
    for row in sheet_range:
        col_list = []
        for col in row:
            col_list.append(col.value)
        data_list += [col_list]

    return data_list


def edit_font(new_ws):
    for rows in new_ws:
        for cell in rows:
            cell.font = Font(size=11, name='游ゴシック')


def write_list(ws, list_2d, start_row, start_col):
    """2次元配列のリストをシートに書き込む"""
    for y, row in enumerate(list_2d):
        for x, cell in enumerate(row):
            ws.cell(row=start_row + y,
                    column=start_col + x,
                    value=list_2d[y][x])


def get_table_download_link(new_wb, input_text, file_count):
    """ダウンロードリンクの生成"""
    wb = save_virtual_workbook(new_wb)
    bs64 = base64.b64encode(wb).decode('UTF-8')
    return f'<a href="data:file/xlsx;base64,{bs64}" download="{input_text}_{file_count}.xlsx">ダウンロード：{input_text}_{file_count}.xlsx</a>'


if __name__ == '__main__':
    main()
